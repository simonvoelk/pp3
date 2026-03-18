"""Generischer XML -> SQLite Importer mit XSD-gestuetzter Typabbildung.

Ziele:
- Direkte XML-Kinder als Haupttabelle importieren.
- Verschachtelte Knoten rekursiv in eigene Tabellen schreiben.
- PK/FK-Auswahl interaktiv vom Benutzer abfragen.
- Beziehungen zwischen Ebenen ueber Link-Tabellen modellieren (m:n-faehig).
- Bei PK-Kollisionen robust weiterlaufen (Duplikate werden uebersprungen).
"""

from __future__ import annotations

import csv
import json
import re
import sqlite3
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple


# ----------------------------
# XSD parsing / type resolution
# ----------------------------

XSD_NS = {"xs": "http://www.w3.org/2001/XMLSchema"}
XSI_SCHEMA_LOCATION = "{http://www.w3.org/2001/XMLSchema-instance}schemaLocation"
XSI_NO_NAMESPACE_SCHEMA_LOCATION = "{http://www.w3.org/2001/XMLSchema-instance}noNamespaceSchemaLocation"
DEFAULT_ZKP_SOURCE_DIR = Path(r"Ed_01_10_2")
ZKP_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xls"}
GENERIC_TARGET_ELEMENT_NAMES = {
    "ITEM",
    "ENTRY",
    "VALUE",
    "ELEMENT",
    "RECORD",
    "ROW",
}


@dataclass
class SimpleTypeInfo:
    """Information zu einem XSD simpleType."""

    name: str
    base: str
    length: Optional[int] = None
    max_length: Optional[int] = None


@dataclass
class ComplexTypeInfo:
    """Information zu einem XSD complexType."""

    name: str
    elements: List[Tuple[str, str]]


@dataclass
class NestedTableState:
    """Laufzeitdaten fuer eine dynamisch erzeugte verschachtelte Tabelle."""

    table_name: str
    columns: List[Tuple[str, str]]
    pk_col: Optional[str]
    pk_autoincrement: bool
    parent_ref_col: Optional[str]
    complex_type_name: Optional[str]
    insert_sql: str
    link_table_name: Optional[str] = None
    link_parent_col: Optional[str] = None
    link_child_col: Optional[str] = None
    link_insert_sql: Optional[str] = None
    inserted_rows: int = 0
    skipped_duplicates: int = 0
    linked_rows: int = 0


@dataclass(frozen=True)
class ImportJob:
    """Ein einzelner Importlauf fuer eine XML-Datei + Ziel-Block."""

    xml_path: Path
    xsd_main: Path
    element_name: str
    complex_type_name: str
    xml_element_name: Optional[str] = None
    parent_element_name: Optional[str] = None


class XsdIndex:
    """Liest XSD-Dateien ein und stellt Typinformationen fuer den Import bereit."""

    BUILTIN_XSD_TYPES = {
        "string",
        "boolean",
        "decimal",
        "float",
        "double",
        "date",
        "dateTime",
        "time",
        "byte",
        "short",
        "int",
        "integer",
        "long",
        "unsignedByte",
        "unsignedShort",
        "unsignedInt",
        "unsignedLong",
    }
    GLOBAL_INLINE_PREFIX = "__el__"
    LOCAL_INLINE_PREFIX = "__ct__"

    def __init__(self) -> None:
        self.simple_types: Dict[str, SimpleTypeInfo] = {}
        self.complex_types: Dict[str, ComplexTypeInfo] = {}
        self.element_to_type: Dict[str, str] = {}
        self.global_elements: Dict[str, ET.Element] = {}
        self._building_complex_types: set[str] = set()

    def load_xsd(self, xsd_path: Path) -> None:
        """Fuegt eine XSD-Datei dem Index hinzu."""
        tree = ET.parse(xsd_path)
        root = tree.getroot()

        self._index_global_element_nodes(root)
        self._index_simple_types(root)
        self._index_complex_types(root)
        self._index_global_elements(root)

    def _index_global_element_nodes(self, root: ET.Element) -> None:
        for element in root.findall("./xs:element", XSD_NS):
            element_name = element.get("name")
            if not element_name:
                continue
            self.global_elements[element_name] = element

    def _index_simple_types(self, root: ET.Element) -> None:
        for simple_type in root.findall(".//xs:simpleType", XSD_NS):
            name = simple_type.get("name")
            if not name:
                continue

            restriction = simple_type.find("xs:restriction", XSD_NS)
            if restriction is None:
                continue

            base = restriction.get("base")
            if not base:
                continue

            length_el = restriction.find("xs:length", XSD_NS)
            maxlen_el = restriction.find("xs:maxLength", XSD_NS)
            length = int(length_el.get("value")) if length_el is not None and length_el.get("value") else None
            max_length = int(maxlen_el.get("value")) if maxlen_el is not None and maxlen_el.get("value") else None

            self.simple_types[name] = SimpleTypeInfo(
                name=name,
                base=base,
                length=length,
                max_length=max_length,
            )

    def _index_complex_types(self, root: ET.Element) -> None:
        for complex_type in root.findall(".//xs:complexType", XSD_NS):
            name = complex_type.get("name")
            if not name:
                continue
            self._build_complex_type(name, complex_type)

    def _index_global_elements(self, root: ET.Element) -> None:
        for element in root.findall("./xs:element", XSD_NS):
            element_name = element.get("name")
            element_type = self._resolve_element_type(element, owner_type_name=None)
            if element_name and element_type:
                self.element_to_type[element_name] = element_type

    @classmethod
    def global_inline_type_name(cls, element_name: str) -> str:
        return f"{cls.GLOBAL_INLINE_PREFIX}{element_name}"

    @classmethod
    def local_inline_type_name(cls, owner_type_name: str, element_name: str) -> str:
        return f"{cls.LOCAL_INLINE_PREFIX}{owner_type_name}__{element_name}"

    def _resolve_element_type(self, element: ET.Element, owner_type_name: Optional[str]) -> Optional[str]:
        element_type = element.get("type")
        if element_type:
            return self.strip_prefix(element_type)

        element_ref = element.get("ref")
        if element_ref:
            ref_name = self.strip_prefix(element_ref)
            ref_element = self.global_elements.get(ref_name)
            if ref_element is None:
                return None
            return self._resolve_element_type(ref_element, owner_type_name=None)

        complex_type = element.find("xs:complexType", XSD_NS)
        if complex_type is None:
            return None

        element_name = element.get("name")
        if not element_name:
            return None

        if owner_type_name:
            type_name = self.local_inline_type_name(owner_type_name, element_name)
        else:
            type_name = self.global_inline_type_name(element_name)

        self._build_complex_type(type_name, complex_type)
        return type_name

    def _build_complex_type(self, type_name: str, complex_type: ET.Element) -> None:
        if type_name in self.complex_types or type_name in self._building_complex_types:
            return

        self._building_complex_types.add(type_name)
        try:
            elements: List[Tuple[str, str]] = []
            seen_cols = set()

            def append_col(col_name: str, col_type: str) -> None:
                if col_name in seen_cols:
                    return
                seen_cols.add(col_name)
                elements.append((col_name, col_type))

            seq = complex_type.find("xs:sequence", XSD_NS)
            if seq is not None:
                for child in seq.findall("xs:element", XSD_NS):
                    child_name = child.get("name")
                    if not child_name:
                        ref_name = child.get("ref")
                        child_name = self.strip_prefix(ref_name) if ref_name else None
                    if not child_name:
                        continue

                    child_type = self._resolve_element_type(child, owner_type_name=type_name)
                    if not child_type:
                        continue
                    append_col(child_name, child_type)

            # Unterstuetzt simpleContent mit Attributen, z. B. <Error code="...">text</Error>.
            extension = complex_type.find("xs:simpleContent/xs:extension", XSD_NS)
            if extension is not None:
                base = extension.get("base", "xs:string")
                append_col("value", base)
                for attr in extension.findall("xs:attribute", XSD_NS):
                    attr_name = attr.get("name")
                    if not attr_name:
                        continue
                    attr_type = attr.get("type", "xs:string")
                    append_col(attr_name, attr_type)

            # Unterstuetzt Attribute direkt am complexType.
            for attr in complex_type.findall("xs:attribute", XSD_NS):
                attr_name = attr.get("name")
                if not attr_name:
                    continue
                attr_type = attr.get("type", "xs:string")
                append_col(attr_name, attr_type)

            self.complex_types[type_name] = ComplexTypeInfo(name=type_name, elements=elements)
        finally:
            self._building_complex_types.remove(type_name)

    @staticmethod
    def strip_prefix(qname: str) -> str:
        """Extrahiert den lokalen Namen aus QName-Formaten wie tns:Foo."""
        return qname.split(":", 1)[1] if ":" in qname else qname

    def get_child_element_type(self, complex_type_name: Optional[str], child_name: str) -> Optional[str]:
        """Liefert den XSD-Typ eines Child-Elements innerhalb eines complexType."""
        if not complex_type_name:
            return None

        complex_type = self.complex_types.get(complex_type_name)
        if not complex_type:
            return None

        for element_name, element_type in complex_type.elements:
            if element_name == child_name:
                return element_type

        return None

    def resolve_to_builtin_base(self, xsd_type_qname: str, depth: int = 0) -> Tuple[str, Optional[int]]:
        """Loest benutzerdefinierte XSD-Typen auf Builtin-Typ + Laengenhinweis auf."""
        if depth > 50:
            raise ValueError(f"Type recursion too deep while resolving: {xsd_type_qname}")

        candidate = xsd_type_qname

        if candidate.startswith("xs:"):
            return self.strip_prefix(candidate), None

        if ":" not in candidate and candidate in self.BUILTIN_XSD_TYPES:
            return candidate, None

        local_name = self.strip_prefix(candidate)
        simple_type = self.simple_types.get(local_name)
        if simple_type:
            base = simple_type.base
            length_hint = simple_type.length or simple_type.max_length

            if base.startswith("tns:") or (":" in base and not base.startswith("xs:")):
                inner_builtin, inner_len = self.resolve_to_builtin_base(base, depth + 1)
                return inner_builtin, length_hint or inner_len

            if base.startswith("xs:"):
                return self.strip_prefix(base), length_hint

            return self.strip_prefix(base), length_hint

        # Fallback: Unbekannter Typname wird als lokaler Name behandelt.
        return local_name, None


# ----------------------------
# SQL helpers
# ----------------------------


def map_xsd_builtin_to_sql(builtin: str, length_hint: Optional[int]) -> str:
    """Mappt XSD-Builtin-Typen auf SQLite-kompatible Datentypen."""
    if builtin in ("string", "normalizedString", "token"):
        return f"VARCHAR({length_hint})" if length_hint else "TEXT"

    if builtin == "boolean":
        return "BOOLEAN"

    if builtin == "decimal":
        return "NUMERIC"

    if builtin in ("float", "double"):
        return "REAL"

    if builtin in (
        "byte",
        "short",
        "int",
        "integer",
        "long",
        "unsignedByte",
        "unsignedShort",
        "unsignedInt",
        "unsignedLong",
    ):
        return "INTEGER"

    if builtin == "dateTime":
        return "TIMESTAMP"
    if builtin == "date":
        return "DATE"
    if builtin == "time":
        return "TIME"

    return "TEXT"


def create_table_sql(
    table_name: str,
    columns: List[Tuple[str, str]],
    pk_col: Optional[str] = None,
    pk_autoincrement: bool = False,
    fks: Optional[List[Tuple[str, str, str]]] = None,
) -> str:
    """Erzeugt CREATE TABLE DDL inklusive optionalem PK/FK."""
    col_defs: List[str] = []
    for column, sql_type in columns:
        if pk_col and pk_autoincrement and column == pk_col:
            col_defs.append(f'"{column}" INTEGER PRIMARY KEY AUTOINCREMENT')
        else:
            col_defs.append(f'"{column}" {sql_type}')

    constraints: List[str] = []

    if pk_col and not pk_autoincrement:
        constraints.append(f'PRIMARY KEY("{pk_col}")')

    if fks:
        for column, ref_table, ref_col in fks:
            constraints.append(f'FOREIGN KEY("{column}") REFERENCES "{ref_table}"("{ref_col}")')

    return (
        f'CREATE TABLE IF NOT EXISTS "{table_name}" (\n  '
        + ",\n  ".join(col_defs + constraints)
        + "\n);"
    )


def create_link_table_sql(
    *,
    link_table_name: str,
    parent_col: str,
    parent_type: str,
    parent_table: str,
    parent_ref_col: str,
    child_col: str,
    child_type: str,
    child_table: str,
    child_ref_col: str,
) -> str:
    """Erzeugt DDL fuer eine m:n-Linktabelle zwischen Parent und Child."""
    return (
        f'CREATE TABLE IF NOT EXISTS "{link_table_name}" (\n'
        f'  "{parent_col}" {parent_type} NOT NULL,\n'
        f'  "{child_col}" {child_type} NOT NULL,\n'
        f'  PRIMARY KEY("{parent_col}", "{child_col}"),\n'
        f'  FOREIGN KEY("{parent_col}") REFERENCES "{parent_table}"("{parent_ref_col}"),\n'
        f'  FOREIGN KEY("{child_col}") REFERENCES "{child_table}"("{child_ref_col}")\n'
        ");"
    )


def build_insert_sql(table_name: str, col_names: List[str]) -> str:
    """Erzeugt ein parametrisiertes INSERT-Statement."""
    placeholders = ", ".join(["?"] * len(col_names))
    cols_sql = ", ".join(f'"{col}"' for col in col_names)
    return f'INSERT INTO "{table_name}" ({cols_sql}) VALUES ({placeholders})'


def build_link_insert_sql(table_name: str, parent_col: str, child_col: str) -> str:
    """Erzeugt INSERT OR IGNORE fuer eine Linktabelle."""
    return (
        f'INSERT OR IGNORE INTO "{table_name}" ("{parent_col}", "{child_col}") '
        "VALUES (?, ?)"
    )


def ensure_table_columns(
    conn: sqlite3.Connection,
    table_name: str,
    required_columns: List[Tuple[str, str]],
) -> None:
    """Ergaenzt fehlende Spalten in bestehenden Tabellen per ALTER TABLE."""
    existing_rows = conn.execute(f'PRAGMA table_info("{table_name}")').fetchall()
    existing_cols = {row[1] for row in existing_rows}  # row[1] = column name

    for column_name, sql_type in required_columns:
        if column_name in existing_cols:
            continue
        conn.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "{column_name}" {sql_type}')
        existing_cols.add(column_name)


COMMON_INDEX_COLUMNS = {
    "SYS_ID",
    "SysID",
    "PV_ID",
    "COB_ID",
    "MUX",
    "NAME",
}


def _safe_index_name(raw_name: str) -> str:
    """Normalisiert Indexnamen fuer SQLite."""
    safe = re.sub(r"[^A-Za-z0-9_]", "_", raw_name)
    return safe[:120] if len(safe) > 120 else safe


def ensure_common_indexes(
    conn: sqlite3.Connection,
    table_name: str,
    columns: List[Tuple[str, str]],
) -> None:
    """Legt hilfreiche Indizes fuer typische Join-Spalten an."""
    for column_name, _sql_type in columns:
        if column_name not in COMMON_INDEX_COLUMNS:
            continue
        index_name = _safe_index_name(f"idx_{table_name}_{column_name}")
        conn.execute(
            f'CREATE INDEX IF NOT EXISTS "{index_name}" '
            f'ON "{table_name}"("{column_name}")'
        )


INT_PATTERN = re.compile(r"^[+-]?\d+$")
REAL_PATTERN = re.compile(r"^[+-]?(\d+\.\d*|\d*\.\d+)([eE][+-]?\d+)?$")
BOOL_WORDS = {"true", "false"}


def infer_sql_type_from_text(raw_value: Optional[str]) -> str:
    """Leitet SQL-Typ aus einem Textwert ab (Fallback ohne XSD)."""
    if raw_value is None:
        return "TEXT"

    value = raw_value.strip()
    if not value:
        return "TEXT"

    lower = value.lower()
    if lower in BOOL_WORDS:
        return "BOOLEAN"

    if INT_PATTERN.match(value):
        return "INTEGER"

    if REAL_PATTERN.match(value):
        return "REAL"

    return "TEXT"


def _normalize_zkp(raw_value: str) -> str:
    """Normalisiert ZKP-Text, damit er einzeilig gespeichert wird."""
    return " ".join(raw_value.split())


def _normalize_sys_id(raw_value: object) -> str:
    """Normalisiert SYS_ID-Werte aus XML/Excel fuer robuste Vergleiche."""
    return str(raw_value).strip() if raw_value is not None else ""


def _canonicalize_sys_id_for_zkp(sys_id: str) -> str:
    """Bildet SYS_ID auf eine kanonische Form ab (letztes Segment -> 0).

    Beispiel:
    0070.081006.010.010.2 -> 0070.081006.010.010.0
    """
    value = _normalize_sys_id(sys_id)
    match = re.match(r"^(.*\.)(\d+)$", value)
    if not match:
        return value
    return f"{match.group(1)}0"


def _build_canonical_zkp_mapping(mapping: Dict[str, str]) -> Dict[str, str]:
    """Erzeugt SYS_ID(.x)-unabhaengige Zuordnung fuer Fallback-Matching.

    Bei Konflikten (gleicher kanonischer Key mit unterschiedlichen ZKPs)
    wird der Key verworfen, um falsche Zuweisungen zu vermeiden.
    """
    canonical: Dict[str, str] = {}
    conflicts: set[str] = set()

    for sys_id, zkp in mapping.items():
        key = _canonicalize_sys_id_for_zkp(sys_id)
        existing = canonical.get(key)
        if existing is None:
            canonical[key] = zkp
            continue
        if existing != zkp:
            conflicts.add(key)

    for key in conflicts:
        canonical.pop(key, None)
    return canonical


def _sys_id_prefix3(sys_id: str) -> str:
    """Liefert den SYS_ID-Praefix aus den ersten 3 Gruppen."""
    value = _normalize_sys_id(sys_id)
    if not value:
        return ""
    parts = value.split(".")
    if len(parts) < 3:
        return ""
    return ".".join(parts[:3])


def _build_prefix3_zkp_mapping(mapping: Dict[str, str]) -> Dict[str, str]:
    """Erzeugt konfliktfreie ZKP-Zuordnung ueber SYS_ID-Praefix (erste 3 Gruppen)."""
    prefix_map: Dict[str, str] = {}
    conflicts: set[str] = set()

    for sys_id, zkp in mapping.items():
        prefix = _sys_id_prefix3(sys_id)
        if not prefix:
            continue
        existing = prefix_map.get(prefix)
        if existing is None:
            prefix_map[prefix] = zkp
            continue
        if existing != zkp:
            conflicts.add(prefix)

    for prefix in conflicts:
        prefix_map.pop(prefix, None)
    return prefix_map


def _lookup_zkp_for_sys_id(
    sys_id: str,
    exact_mapping: Dict[str, str],
    canonical_mapping: Dict[str, str],
    prefix3_mapping: Dict[str, str],
) -> Optional[str]:
    """Sucht ZKP fuer eine SYS_ID mit Fallback-Kaskade."""
    value = _normalize_sys_id(sys_id)
    if not value:
        return None

    exact = exact_mapping.get(value)
    if exact:
        return exact

    canonical = canonical_mapping.get(_canonicalize_sys_id_for_zkp(value))
    if canonical:
        return canonical

    return prefix3_mapping.get(_sys_id_prefix3(value))


def _parse_sysid_zkp_rows(rows: List[List[object]]) -> Dict[str, str]:
    """Extrahiert SYS_ID -> ZKP aus tabellarischen Zeilen."""
    def normalize_header(cell: object) -> str:
        return re.sub(r"[^A-Z0-9]", "", str(cell).upper()) if cell is not None else ""

    def score_sys_header(cell: object) -> int:
        norm = normalize_header(cell)
        if norm == "SYSID":
            return 4
        if norm.startswith("SYSID"):
            return 3
        if "SYSID" in norm:
            return 2
        if norm.startswith("SYS") and "ID" in norm:
            return 1
        return 0

    def score_zkp_header(cell: object) -> int:
        norm = normalize_header(cell)
        if norm == "ZKP":
            return 4
        if norm.startswith("ZKP") or norm.endswith("ZKP"):
            return 3
        if "ZKP" in norm:
            return 2
        return 0

    def looks_like_sys_id(value: str) -> bool:
        parts = value.split(".")
        if len(parts) < 3:
            return False
        return all(part.isdigit() for part in parts)

    def is_valid_zkp_cell(raw_value: object) -> bool:
        if raw_value is None:
            return False
        text = str(raw_value).strip()
        if not text:
            return False
        upper = text.upper()
        return upper not in {"#N/A", "N/A", "NA", "NULL", "-"}

    def count_sys_hits(start_idx: int, col_idx: int, window: int = 250) -> int:
        end = min(len(rows), start_idx + 1 + window)
        hits = 0
        for r in rows[start_idx + 1 : end]:
            if col_idx >= len(r):
                continue
            value = _normalize_sys_id(r[col_idx])
            if looks_like_sys_id(value):
                hits += 1
        return hits

    def count_zkp_hits(start_idx: int, col_idx: int, window: int = 250) -> int:
        end = min(len(rows), start_idx + 1 + window)
        hits = 0
        for r in rows[start_idx + 1 : end]:
            if col_idx >= len(r):
                continue
            if is_valid_zkp_cell(r[col_idx]):
                hits += 1
        return hits

    header_idx: Optional[int] = None
    sys_col = -1
    zkp_col = -1
    best_score = -1

    for idx, row in enumerate(rows):
        sys_candidates = [(col_idx, score_sys_header(cell)) for col_idx, cell in enumerate(row)]
        sys_candidates = [item for item in sys_candidates if item[1] > 0]
        zkp_candidates = [(col_idx, score_zkp_header(cell)) for col_idx, cell in enumerate(row)]
        zkp_candidates = [item for item in zkp_candidates if item[1] > 0]
        if not sys_candidates or not zkp_candidates:
            continue

        sys_ranked = [
            (col_idx, hdr_score, count_sys_hits(idx, col_idx))
            for col_idx, hdr_score in sys_candidates
        ]
        sys_ranked.sort(key=lambda item: (item[2], item[1]), reverse=True)
        best_sys_col, best_sys_hdr_score, best_sys_hits = sys_ranked[0]
        if best_sys_hits == 0:
            continue

        zkp_ranked = [
            (col_idx, hdr_score, count_zkp_hits(idx, col_idx))
            for col_idx, hdr_score in zkp_candidates
        ]
        zkp_ranked.sort(key=lambda item: (item[2], item[1]), reverse=True)
        best_zkp_col, best_zkp_hdr_score, best_zkp_hits = zkp_ranked[0]

        score = best_sys_hits * 1000 + best_zkp_hits * 10 + best_sys_hdr_score + best_zkp_hdr_score
        if score > best_score:
            best_score = score
            header_idx = idx
            sys_col = best_sys_col
            zkp_col = best_zkp_col

    if header_idx is None or sys_col < 0 or zkp_col < 0:
        return {}

    mapping: Dict[str, str] = {}
    for row in rows[header_idx + 1 :]:
        if sys_col >= len(row):
            continue
        sys_id = _normalize_sys_id(row[sys_col])
        if not sys_id:
            continue
        if not looks_like_sys_id(sys_id):
            continue

        raw_cell = row[zkp_col] if zkp_col < len(row) else None
        if not is_valid_zkp_cell(raw_cell):
            continue
        raw_zkp = str(raw_cell).strip()
        if not raw_zkp:
            continue

        zkp = _normalize_zkp(raw_zkp)
        if zkp and sys_id not in mapping:
            mapping[sys_id] = zkp

    return mapping


def _load_sysid_zkp_from_csv(csv_path: Path) -> Dict[str, str]:
    """Laedt SYS_ID -> ZKP aus einer CSV-Datei."""
    with csv_path.open("r", encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.reader(fh, delimiter=";"))
    return _parse_sysid_zkp_rows([[cell for cell in row] for row in rows])


def _load_sysid_zkp_from_excel(excel_path: Path) -> Dict[str, str]:
    """Laedt SYS_ID -> ZKP aus einer Excel-Datei (xlsx/xlsm/xls)."""
    suffix = excel_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        try:
            import openpyxl  # type: ignore
        except ModuleNotFoundError:
            return {}

        workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        mapping: Dict[str, str] = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            part = _parse_sysid_zkp_rows(rows)
            for sys_id, zkp in part.items():
                if sys_id not in mapping:
                    mapping[sys_id] = zkp
        return mapping

    if suffix == ".xls":
        try:
            import xlrd  # type: ignore
        except ModuleNotFoundError:
            return {}

        workbook = xlrd.open_workbook(excel_path)
        mapping = {}
        for sheet in workbook.sheets():
            rows = [sheet.row_values(idx) for idx in range(sheet.nrows)]
            part = _parse_sysid_zkp_rows(rows)
            for sys_id, zkp in part.items():
                if sys_id not in mapping:
                    mapping[sys_id] = zkp
        return mapping

    return {}


def _collect_zkp_source_files(source_path: Path) -> List[Path]:
    """Sammelt zu beruecksichtigende ZKP-Quelldateien."""
    if source_path.is_dir():
        return sorted(
            [
                path
                for path in source_path.iterdir()
                if path.is_file()
                and path.suffix.lower() in ZKP_EXCEL_SUFFIXES
                and not path.name.startswith("~$")
            ],
            key=lambda path: path.name.lower(),
        )

    if source_path.is_file():
        if source_path.suffix.lower() in ZKP_EXCEL_SUFFIXES or source_path.suffix.lower() == ".csv":
            return [source_path]
    return []


def load_sysid_zkp_mapping(source_path: Path = DEFAULT_ZKP_SOURCE_DIR) -> Tuple[Dict[str, str], List[str]]:
    """Laedt SYS_ID -> ZKP aus allen Excel-Quellen (mit CSV-Fallback pro Datei)."""
    mapping: Dict[str, str] = {}
    used_sources: List[str] = []

    for source_file in _collect_zkp_source_files(source_path):
        part: Dict[str, str] = {}
        used_file: Optional[Path] = None

        try:
            if source_file.suffix.lower() == ".csv":
                part = _load_sysid_zkp_from_csv(source_file)
                used_file = source_file
            else:
                part = _load_sysid_zkp_from_excel(source_file)
                used_file = source_file if part else None
        except Exception:
            part = {}
            used_file = None

        # Wie bisher: bei Excel-Fehlern/Funktionsluecken auf gleichnamige CSV-Datei fallen.
        if not part and source_file.suffix.lower() in ZKP_EXCEL_SUFFIXES:
            csv_fallback = source_file.with_suffix(".csv")
            if csv_fallback.exists():
                try:
                    part = _load_sysid_zkp_from_csv(csv_fallback)
                    used_file = csv_fallback if part else None
                except Exception:
                    part = {}
                    used_file = None

        if not part:
            continue

        for sys_id, zkp in part.items():
            if sys_id not in mapping:
                mapping[sys_id] = zkp

        if used_file is not None:
            used_sources.append(str(used_file))

    return mapping, used_sources


def _get_sysid_columns_by_table(conn: sqlite3.Connection) -> List[Tuple[str, List[str]]]:
    """Liefert alle Tabellen plus erkannte SYS_ID-Spalten."""
    tables = [
        row[0]
        for row in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
        )
    ]

    result: List[Tuple[str, List[str]]] = []
    for table in tables:
        table_info = conn.execute(f'PRAGMA table_info("{table}")').fetchall()
        ordered_cols = [row[1] for row in table_info]
        sys_cols = [
            col
            for col in ordered_cols
            if "SYS_ID" in col.upper() or col == "SysID"
        ]
        if sys_cols:
            result.append((table, sys_cols))
    return result


def sync_zkp_to_database(sqlite_path: Path, source_path: Path = DEFAULT_ZKP_SOURCE_DIR) -> Dict[str, int]:
    """Synchronisiert ZKP-Werte anhand SYS_ID in allen passenden Tabellen."""
    mapping, sources_used = load_sysid_zkp_mapping(source_path)
    if not mapping:
        print(
            f"Hinweis: Keine SYS_ID->ZKP-Zuordnung geladen "
            f'(erwartet Excel-Dateien in "{source_path}" inkl. CSV-Fallback pro Datei).'
        )
        return {}

    source_info = f"{len(sources_used)} Datei(en)"
    if sources_used:
        preview = ", ".join(Path(path).name for path in sources_used[:3])
        if len(sources_used) > 3:
            preview += ", ..."
        source_info += f" ({preview})"

    canonical_mapping = _build_canonical_zkp_mapping(mapping)
    prefix3_mapping = _build_prefix3_zkp_mapping(mapping)
    print(
        f'ZKP-Sync: Quellen {source_info}, '
        f'Eintraege exakt: {len(mapping)}, kanonisch: {len(canonical_mapping)}, '
        f'Prefix3: {len(prefix3_mapping)}'
    )

    stats: Dict[str, int] = {}
    with sqlite3.connect(sqlite_path) as conn:
        # Robust gegen Umgebungen, in denen Datei-Journale fehlschlagen.
        conn.execute("PRAGMA journal_mode=MEMORY")
        conn.execute("PRAGMA synchronous=NORMAL")
        tables = _get_sysid_columns_by_table(conn)
        if not tables:
            return stats

        for table_name, sys_cols in tables:
            ensure_table_columns(conn, table_name, [("ZKP", "TEXT")])

            conn.execute(f'UPDATE "{table_name}" SET "ZKP" = NULL')

            before = conn.total_changes
            select_cols_sql = ", ".join(f'"{col}"' for col in sys_cols)
            row_data = conn.execute(
                f'SELECT rowid, {select_cols_sql} FROM "{table_name}"'
            ).fetchall()
            fallback_updates: List[Tuple[str, int]] = []
            for row in row_data:
                rowid = row[0]
                candidate_sys_ids = row[1:]
                zkp: Optional[str] = None
                for raw_sys_id in candidate_sys_ids:
                    sys_id = _normalize_sys_id(raw_sys_id)
                    if not sys_id:
                        continue
                    zkp = _lookup_zkp_for_sys_id(
                        sys_id,
                        mapping,
                        canonical_mapping,
                        prefix3_mapping,
                    )
                    if zkp:
                        break
                if zkp:
                    fallback_updates.append((zkp, rowid))

            if fallback_updates:
                conn.executemany(
                    f'UPDATE "{table_name}" SET "ZKP" = ? WHERE rowid = ?',
                    fallback_updates,
                )

            stats[table_name] = conn.total_changes - before

        conn.commit()

    return stats


def merge_columns(*groups: List[Tuple[str, str]]) -> List[Tuple[str, str]]:
    """Fuegt mehrere Spaltenlisten zusammen, ohne doppelte Spaltennamen."""
    merged: List[Tuple[str, str]] = []
    seen = set()

    for group in groups:
        for column_name, sql_type in group:
            if column_name in seen:
                continue
            merged.append((column_name, sql_type))
            seen.add(column_name)

    return merged


def prompt_choice(prompt: str, options: List[str]) -> Optional[str]:
    """Fragt eine Auswahl interaktiv ab. `0` bedeutet keine Auswahl."""
    if not options:
        return None

    print(prompt)
    for idx, option in enumerate(options, start=1):
        print(f"  {idx}. {option}")
    print("  0. (keins)")

    while True:
        raw = input("Auswahl: ").strip()
        if raw.isdigit():
            index = int(raw)
            if index == 0:
                return None
            if 1 <= index <= len(options):
                return options[index - 1]
        print("Ungueltig. Bitte Zahl eingeben.")


def prompt_required_choice(prompt: str, options: List[str]) -> str:
    """Wie prompt_choice, aber ohne Option auf leere Auswahl."""
    if not options:
        raise ValueError("Keine Optionen fuer erforderliche Auswahl vorhanden.")

    print(prompt)
    for idx, option in enumerate(options, start=1):
        print(f"  {idx}. {option}")

    while True:
        raw = input("Auswahl: ").strip()
        if raw.isdigit():
            index = int(raw)
            if 1 <= index <= len(options):
                return options[index - 1]
        print("Ungueltig. Bitte Zahl eingeben.")


PK_AUTO_PREFIX = "__AUTO__:"


def encode_auto_pk(column_name: str) -> str:
    """Kodiert eine PK-Auswahl als autoinkrementierende ID-Spalte."""
    return f"{PK_AUTO_PREFIX}{column_name}"


def decode_pk_choice(choice: str) -> Tuple[str, bool]:
    """Dekodiert gespeicherte PK-Auswahl in (Spaltenname, autoincrement?)."""
    if choice.startswith(PK_AUTO_PREFIX):
        return choice[len(PK_AUTO_PREFIX) :], True
    return choice, False


def suggest_auto_pk_column(base_name: str, existing_columns: List[str]) -> str:
    """Erzeugt einen kollisionsfreien Spaltennamen fuer eine Auto-ID."""
    taken = set(existing_columns)
    candidate = base_name
    suffix = 1
    while candidate in taken:
        candidate = f"{base_name}_{suffix}"
        suffix += 1
    return candidate


def prompt_pk_choice(prompt: str, options: List[str], auto_column_name: str) -> str:
    """Fragt PK-Auswahl ab und bietet zusaetzlich AUTOINCREMENT an."""
    if not options:
        raise ValueError("Keine Optionen fuer erforderliche Auswahl vorhanden.")

    print(prompt)
    for idx, option in enumerate(options, start=1):
        print(f"  {idx}. {option}")
    auto_idx = len(options) + 1
    print(f"  {auto_idx}. (AUTOINCREMENT, neue ID-Spalte: {auto_column_name})")

    while True:
        raw = input("Auswahl: ").strip()
        if raw.isdigit():
            index = int(raw)
            if 1 <= index <= len(options):
                return options[index - 1]
            if index == auto_idx:
                return encode_auto_pk(auto_column_name)
        print("Ungueltig. Bitte Zahl eingeben.")


def prompt_text(prompt: str, default: Optional[str] = None) -> str:
    """Fragt einen freien Texteingabewert ab."""
    while True:
        suffix = f' [{default}]' if default else ""
        raw = input(f"{prompt}{suffix}: ").strip()
        if raw:
            return raw
        if default is not None:
            return default
        print("Ungueltig. Bitte einen Wert eingeben.")


def prompt_multi_file_selection(file_names: List[str]) -> List[str]:
    """Fragt eine Mehrfachauswahl von XML-Dateien ab."""
    if not file_names:
        return []

    print("Welche XML-Dateien sollen in die DB ueberfuehrt werden?")
    print("Eingabe: 'all' oder Nummern kommasepariert, z. B. 1,3,7")
    for idx, name in enumerate(file_names, start=1):
        print(f"  {idx}. {name}")

    while True:
        raw = input("Auswahl: ").strip().lower()
        if raw in {"all", "a", "*"}:
            return file_names

        parts = [p.strip() for p in raw.split(",") if p.strip()]
        if not parts:
            print("Ungueltig. Bitte 'all' oder Nummern angeben.")
            continue

        picked_indices: List[int] = []
        valid = True
        for part in parts:
            if not part.isdigit():
                valid = False
                break
            idx = int(part)
            if not (1 <= idx <= len(file_names)):
                valid = False
                break
            picked_indices.append(idx)

        if not valid:
            print("Ungueltig. Bitte gueltige Nummern angeben.")
            continue

        # Duplikate entfernen, Reihenfolge beibehalten
        seen = set()
        selected: List[str] = []
        for idx in picked_indices:
            name = file_names[idx - 1]
            if name in seen:
                continue
            seen.add(name)
            selected.append(name)

        if selected:
            return selected

        print("Ungueltig. Bitte mindestens eine Datei auswaehlen.")


class KeySelectionManager:
    """Persistiert PK-Auswahlen fuer Top- und Nested-Tabellen in JSON."""

    def __init__(self, config_path: Path, mode: str) -> None:
        self.config_path = config_path
        self.mode = mode  # "configure" or "update"
        self.data: Dict[str, object] = {
            "version": 1,
            "imports": {},
            "update_files": [],
            "xsd_overrides": {},
            "manual_targets": {},
        }
        self.dirty = False

        if config_path.exists():
            self.data = json.loads(config_path.read_text(encoding="utf-8"))
            self.data.setdefault("imports", {})
            self.data.setdefault("update_files", [])
            self.data.setdefault("xsd_overrides", {})
            self.data.setdefault("manual_targets", {})

    @property
    def imports(self) -> Dict[str, object]:
        imports_obj = self.data.setdefault("imports", {})
        if not isinstance(imports_obj, dict):
            raise ValueError("Ungueltige Key-Konfiguration: 'imports' muss ein Objekt sein.")
        return imports_obj

    def _get_import_cfg(self, import_key: str) -> Dict[str, object]:
        raw = self.imports.get(import_key)
        if isinstance(raw, dict):
            raw.setdefault("nested_pk", {})
            return raw

        cfg: Dict[str, object] = {"top_pk": None, "nested_pk": {}}
        self.imports[import_key] = cfg
        return cfg

    @staticmethod
    def _require_valid_choice(stored: Optional[str], options: List[str], context: str) -> str:
        if stored and stored in options:
            return stored
        if stored and stored.startswith(PK_AUTO_PREFIX):
            auto_col, _ = decode_pk_choice(stored)
            if auto_col:
                return stored
        if stored and stored not in options:
            raise ValueError(f"Gespeicherter Key '{stored}' fuer {context} ist nicht mehr gueltig.")
        raise ValueError(f"Kein gespeicherter Key fuer {context}.")

    def choose_top_pk(
        self,
        import_key: str,
        options: List[str],
        prompt: str,
        auto_column_name: str,
    ) -> str:
        cfg = self._get_import_cfg(import_key)
        stored = cfg.get("top_pk")

        if self.mode == "update":
            return self._require_valid_choice(stored if isinstance(stored, str) else None, options, import_key)

        selected = prompt_pk_choice(prompt, options, auto_column_name)
        cfg["top_pk"] = selected
        self.dirty = True
        return selected

    def choose_nested_pk(
        self,
        import_key: str,
        state_key: str,
        options: List[str],
        prompt: str,
        auto_column_name: str,
    ) -> str:
        cfg = self._get_import_cfg(import_key)
        nested = cfg.get("nested_pk")
        if not isinstance(nested, dict):
            nested = {}
            cfg["nested_pk"] = nested

        stored = nested.get(state_key)
        if self.mode == "update":
            return self._require_valid_choice(
                stored if isinstance(stored, str) else None,
                options,
                f"{import_key}:{state_key}",
            )

        selected = prompt_pk_choice(prompt, options, auto_column_name)
        nested[state_key] = selected
        self.dirty = True
        return selected

    def save(self) -> None:
        if not self.dirty:
            return
        self.config_path.parent.mkdir(parents=True, exist_ok=True)
        self.config_path.write_text(json.dumps(self.data, indent=2, ensure_ascii=True), encoding="utf-8")
        self.dirty = False

    def get_update_files(self) -> List[str]:
        raw = self.data.get("update_files", [])
        if not isinstance(raw, list):
            return []
        return [item for item in raw if isinstance(item, str)]

    def set_update_files(self, file_names: List[str]) -> None:
        unique: List[str] = []
        seen = set()
        for name in file_names:
            if name in seen:
                continue
            seen.add(name)
            unique.append(name)
        self.data["update_files"] = unique
        self.dirty = True

    def get_xsd_override(self, xml_file_name: str) -> Optional[str]:
        raw = self.data.get("xsd_overrides", {})
        if not isinstance(raw, dict):
            return None
        value = raw.get(xml_file_name)
        return value if isinstance(value, str) else None

    def set_xsd_override(self, xml_file_name: str, xsd_file_name: str) -> None:
        raw = self.data.get("xsd_overrides")
        if not isinstance(raw, dict):
            raw = {}
            self.data["xsd_overrides"] = raw
        raw[xml_file_name] = xsd_file_name
        self.dirty = True

    def get_manual_target(self, xml_file_name: str) -> Optional[Tuple[str, str]]:
        raw = self.data.get("manual_targets", {})
        if not isinstance(raw, dict):
            return None
        entry = raw.get(xml_file_name)
        if isinstance(entry, dict):
            element_name = entry.get("element_name")
            complex_type_name = entry.get("complex_type_name")
            if isinstance(element_name, str) and isinstance(complex_type_name, str):
                return (element_name, complex_type_name)

        # Toleriert das zwischenzeitliche Listenformat und nimmt den ersten Eintrag.
        if isinstance(entry, list):
            for item in entry:
                if not isinstance(item, dict):
                    continue
                element_name = item.get("element_name")
                complex_type_name = item.get("complex_type_name")
                if isinstance(element_name, str) and isinstance(complex_type_name, str):
                    return (element_name, complex_type_name)

        return None

    def set_manual_target(self, xml_file_name: str, element_name: str, complex_type_name: str) -> None:
        raw = self.data.get("manual_targets")
        if not isinstance(raw, dict):
            raw = {}
            self.data["manual_targets"] = raw
        raw[xml_file_name] = {
            "element_name": element_name,
            "complex_type_name": complex_type_name,
        }
        self.dirty = True


# ----------------------------
# XML helpers
# ----------------------------


def localname(tag: str) -> str:
    """Entfernt ggf. den Namespace-Prefix aus einem XML-Tag."""
    return tag.split("}", 1)[1] if tag.startswith("{") else tag


def has_element_children(element: ET.Element) -> bool:
    """True, wenn ein XML-Element weitere XML-Element-Kinder besitzt."""
    return len(list(element)) > 0


def casefold_path_map(paths: List[Path]) -> Dict[str, Path]:
    """Mappt Dateinamen case-insensitiv auf ihren Path."""
    return {path.name.lower(): path for path in paths}


def resolve_xsd_for_xml(xml_path: Path, schema_lookup: Dict[str, Path]) -> Optional[Path]:
    """Liest die zugehoerige XSD aus XML-Header; bei Bedarf Fallback auf Dateinamen."""
    try:
        root = ET.parse(xml_path).getroot()
    except ET.ParseError:
        root = None

    locations: List[str] = []
    if root is not None:
        schema_location = root.attrib.get(XSI_SCHEMA_LOCATION, "").strip()
        no_namespace_location = root.attrib.get(XSI_NO_NAMESPACE_SCHEMA_LOCATION, "").strip()

        if schema_location:
            tokens = schema_location.split()
            # Standardfall: schemaLocation ist paarweise aufgebaut
            # (namespace-uri, xsd-path, namespace-uri, xsd-path, ...)
            if len(tokens) >= 2:
                locations.extend(tokens[1::2])
            elif len(tokens) == 1:
                # Fallback fuer nicht standardkonforme Einzelangabe
                locations.append(tokens[0])

        if no_namespace_location:
            locations.append(no_namespace_location)

    for location in locations:
        xsd_name = Path(location).name.lower()
        if xsd_name in schema_lookup:
            return schema_lookup[xsd_name]

    # Fallback: gleichnamige XSD zur XML (z. B. Foo.xml -> Foo.xsd).
    same_name_xsd = f"{xml_path.stem}.xsd".lower()
    return schema_lookup.get(same_name_xsd)


def discover_import_targets_from_xsd(
    xsd_path: Path,
) -> List[Tuple[str, str, Optional[str], Optional[str]]]:
    """Ermittelt importierbare Ziele aus einer XSD.

    Rueckgabe pro Ziel:
    (tabellenname, complex_type_name, xml_element_name, parent_element_name)
    """
    root = ET.parse(xsd_path).getroot()
    global_elements: Dict[str, ET.Element] = {}
    for element in root.findall("./xs:element", XSD_NS):
        element_name = element.get("name")
        if element_name:
            global_elements[element_name] = element

    named_complex_types: Dict[str, ET.Element] = {}
    for complex_type in root.findall(".//xs:complexType", XSD_NS):
        ct_name = complex_type.get("name")
        if ct_name:
            named_complex_types[ct_name] = complex_type
    known_complex_types: set[str] = set(named_complex_types.keys())

    referenced_globals = set()
    for element in root.findall(".//xs:element", XSD_NS):
        ref_name = element.get("ref")
        if ref_name:
            referenced_globals.add(XsdIndex.strip_prefix(ref_name))

    children_cache: Dict[str, List[Tuple[str, str, str]]] = {}
    building_types: set[str] = set()

    def resolve_element_type(
        element: ET.Element,
        owner_type_name: Optional[str],
    ) -> Optional[str]:
        element_type = element.get("type")
        if element_type:
            return XsdIndex.strip_prefix(element_type)

        element_ref = element.get("ref")
        if element_ref:
            ref_name = XsdIndex.strip_prefix(element_ref)
            ref_element = global_elements.get(ref_name)
            if ref_element is None:
                return None
            return resolve_element_type(ref_element, owner_type_name=None)

        complex_type = element.find("xs:complexType", XSD_NS)
        if complex_type is None:
            return None

        element_name = element.get("name")
        if not element_name:
            return None

        if owner_type_name:
            type_name = XsdIndex.local_inline_type_name(owner_type_name, element_name)
        else:
            type_name = XsdIndex.global_inline_type_name(element_name)
        ensure_children_for_type(type_name, complex_type)
        return type_name

    def ensure_children_for_type(type_name: str, complex_type: ET.Element) -> None:
        if type_name in children_cache or type_name in building_types:
            return

        known_complex_types.add(type_name)
        building_types.add(type_name)
        try:
            entries: List[Tuple[str, str, str]] = []
            seq = complex_type.find("xs:sequence", XSD_NS)
            if seq is not None:
                seq_max_occurs = seq.get("maxOccurs", "1")
                for child in seq.findall("xs:element", XSD_NS):
                    child_name = child.get("name")
                    if not child_name:
                        ref_name = child.get("ref")
                        child_name = XsdIndex.strip_prefix(ref_name) if ref_name else None
                    if not child_name:
                        continue

                    child_type = resolve_element_type(child, owner_type_name=type_name)
                    if not child_type:
                        continue

                    child_max_occurs = child.get("maxOccurs", "1")
                    effective_max = child_max_occurs if child_max_occurs != "1" else seq_max_occurs
                    entries.append((child_name, child_type, effective_max))

            children_cache[type_name] = entries
        finally:
            building_types.remove(type_name)

    def get_children_for_type(type_name: str) -> List[Tuple[str, str, str]]:
        if type_name in children_cache:
            return children_cache[type_name]

        named = named_complex_types.get(type_name)
        if named is not None:
            ensure_children_for_type(type_name, named)
            return children_cache.get(type_name, [])

        return []

    def is_complex_type(type_name: str) -> bool:
        if type_name in known_complex_types:
            return True
        # Triggert lazy Aufloesung fuer benannte Typen.
        _ = get_children_for_type(type_name)
        return type_name in known_complex_types

    def resolve_global_type(element_name: str, element: ET.Element) -> Optional[str]:
        type_name = resolve_element_type(element, owner_type_name=None)
        if type_name:
            return type_name
        # Inline simpleType ohne complexType ist fuer Tabellenziele nicht relevant.
        return None

    targets: List[Tuple[str, str, Optional[str], Optional[str]]] = []
    top_candidates: List[Tuple[str, ET.Element]] = []
    for element_name, element in global_elements.items():
        if element_name not in referenced_globals:
            top_candidates.append((element_name, element))
    if not top_candidates:
        top_candidates = list(global_elements.items())

    for top_name, top_element in top_candidates:
        top_type_name = resolve_global_type(top_name, top_element)
        if not top_type_name:
            continue

        for child_name, child_type_name, child_max_occurs in get_children_for_type(top_type_name):
            if not is_complex_type(child_type_name):
                continue

            if child_max_occurs != "1":
                table_name = child_name
                if child_name.upper() in GENERIC_TARGET_ELEMENT_NAMES:
                    table_name = f"{top_name}__{child_name}"
                targets.append((table_name, child_type_name, child_name, top_name))

            # Zweistufiger Fall: einmaliges Parent mit wiederholbarem Subchild.
            for sub_name, sub_type_name, sub_max_occurs in get_children_for_type(child_type_name):
                if not is_complex_type(sub_type_name):
                    continue
                if sub_max_occurs == "1":
                    continue
                table_name = f"{child_name}_{sub_name}"
                targets.append((table_name, sub_type_name, sub_name, child_name))

    unique: List[Tuple[str, str, Optional[str], Optional[str]]] = []
    seen = set()
    for item in targets:
        if item in seen:
            continue
        unique.append(item)
        seen.add(item)
    return unique


def list_complex_types_in_xsd(xsd_path: Path) -> List[str]:
    """Liefert alle benannten complexType-Namen aus einer XSD."""
    root = ET.parse(xsd_path).getroot()
    names: List[str] = []
    seen = set()
    for ct in root.findall(".//xs:complexType", XSD_NS):
        name = ct.get("name")
        if not name or name in seen:
            continue
        seen.add(name)
        names.append(name)
    return sorted(names, key=str.lower)


def prompt_manual_target_for_xsd(xml_path: Path, xsd_path: Path) -> Optional[Tuple[str, str]]:
    """Erfragt manuell Elementname + complexType fuer eine XML/XSD-Kombination."""
    complex_types = list_complex_types_in_xsd(xsd_path)
    if not complex_types:
        print(f'In "{xsd_path.name}" wurden keine benannten complexType-Eintraege gefunden.')
        return None

    complex_type_name = prompt_choice(
        (
            f'Fuer "{xml_path.name}" wurden in "{xsd_path.name}" keine Importziele erkannt. '
            "Bitte complexType manuell waehlen:"
        ),
        complex_types,
    )
    if not complex_type_name:
        return None

    default_element = (
        complex_type_name[:-5] if complex_type_name.lower().endswith("_type") else complex_type_name
    )
    element_name = prompt_text(
        f'Elementname im XML fuer complexType "{complex_type_name}" eingeben',
        default=default_element,
    )
    return (element_name, complex_type_name)


def discover_jobs(
    project_data_dir: Path,
    schemas_dir: Path,
    key_manager: Optional[KeySelectionManager] = None,
    allow_manual_xsd_selection: bool = False,
    xml_name_filter: Optional[set[str]] = None,
) -> List[ImportJob]:
    """Findet relevante XML-Dateien und dazugehoerige Importziele."""
    xsd_files = [p for p in schemas_dir.glob("*.xsd") if p.is_file()]
    schema_lookup = casefold_path_map(xsd_files)
    xsd_file_names = sorted([p.name for p in xsd_files], key=str.lower)

    xml_files = list_project_data_xml_files(project_data_dir)
    if xml_name_filter is not None:
        xml_files = [p for p in xml_files if p.name in xml_name_filter]

    jobs: List[ImportJob] = []
    seen = set()

    for xml_path in xml_files:
        xsd_path = resolve_xsd_for_xml(xml_path, schema_lookup)

        if not xsd_path and key_manager:
            override_name = key_manager.get_xsd_override(xml_path.name)
            if override_name:
                xsd_path = schema_lookup.get(override_name.lower())

        if not xsd_path and allow_manual_xsd_selection:
            chosen = prompt_choice(
                f'Keine passende XSD fuer "{xml_path.name}" gefunden. Manuell XSD waehlen:',
                xsd_file_names,
            )
            if chosen:
                xsd_path = schema_lookup.get(chosen.lower())
                if xsd_path and key_manager:
                    key_manager.set_xsd_override(xml_path.name, xsd_path.name)

        if not xsd_path:
            continue

        targets = discover_import_targets_from_xsd(xsd_path)
        manual_target = key_manager.get_manual_target(xml_path.name) if key_manager else None
        if manual_target:
            manual_element_name, manual_complex_type_name = manual_target
            mapped_target: Optional[Tuple[str, str, Optional[str], Optional[str]]] = None

            # Legacy-Fix: Fruehere Konfigurationen speicherten generische Namen wie "ITEM".
            # Wenn wir passende automatisch erkannte Targets finden, mappen wir darauf um
            # (z. B. LOCAL_SYS_ID_XML__ITEM statt ITEM), damit keine Tabellenkollision entsteht.
            if manual_element_name.upper() in GENERIC_TARGET_ELEMENT_NAMES:
                for candidate in targets:
                    table_name, complex_type_name, xml_element_name, parent_element_name = candidate
                    if complex_type_name != manual_complex_type_name:
                        continue
                    source_name = xml_element_name or table_name
                    if source_name.upper() == manual_element_name.upper():
                        mapped_target = candidate
                        break

            if mapped_target:
                targets = [mapped_target]
            else:
                targets = [(manual_element_name, manual_complex_type_name, None, None)]

        # Falls zwar eine XSD gefunden wurde, aber daraus keine Importziele ableitbar sind,
        # kann der Nutzer eine andere XSD manuell waehlen.
        if not targets and allow_manual_xsd_selection:
            manual_target = prompt_manual_target_for_xsd(xml_path, xsd_path)
            if manual_target:
                element_name, complex_type_name = manual_target
                targets = [(element_name, complex_type_name, None, None)]
                if key_manager:
                    key_manager.set_manual_target(xml_path.name, element_name, complex_type_name)

        if not targets and allow_manual_xsd_selection:
            while True:
                chosen = prompt_choice(
                    (
                        f'Fuer "{xml_path.name}" wurden in "{xsd_path.name}" keine Importziele gefunden. '
                        "Andere XSD manuell waehlen:"
                    ),
                    xsd_file_names,
                )
                if not chosen:
                    break

                manual_xsd = schema_lookup.get(chosen.lower())
                if not manual_xsd:
                    print("Ungueltige XSD-Auswahl.")
                    continue

                manual_targets = discover_import_targets_from_xsd(manual_xsd)
                if not manual_targets:
                    manual_target = prompt_manual_target_for_xsd(xml_path, manual_xsd)
                    if manual_target:
                        element_name, complex_type_name = manual_target
                        manual_targets = [(element_name, complex_type_name, None, None)]
                        if key_manager:
                            key_manager.set_manual_target(xml_path.name, element_name, complex_type_name)
                    else:
                        print(f'Hinweis: In "{manual_xsd.name}" wurden ebenfalls keine Importziele gefunden.')
                        continue

                xsd_path = manual_xsd
                targets = manual_targets
                if key_manager:
                    key_manager.set_xsd_override(xml_path.name, xsd_path.name)
                break

        if not targets:
            continue

        for table_name, complex_type_name, xml_element_name, parent_element_name in targets:
            key = (
                xml_path.resolve(),
                xsd_path.resolve(),
                table_name,
                complex_type_name,
                xml_element_name,
                parent_element_name,
            )
            if key in seen:
                continue
            seen.add(key)
            jobs.append(
                ImportJob(
                    xml_path=xml_path,
                    xsd_main=xsd_path,
                    element_name=table_name,
                    complex_type_name=complex_type_name,
                    xml_element_name=xml_element_name,
                    parent_element_name=parent_element_name,
                )
            )

    return jobs


def list_project_data_xml_files(project_data_dir: Path) -> List[Path]:
    """Liefert alle XML-Dateien im project_data-Ordner (case-insensitiv)."""
    xml_files = [p for p in project_data_dir.iterdir() if p.is_file() and p.suffix.lower() == ".xml"]
    return sorted(xml_files, key=lambda p: p.name.lower())


def build_import_key(project_data_dir: Path, job: ImportJob) -> str:
    """Stabiler Key fuer gespeicherte Key-Auswahlen."""
    rel = job.xml_path.resolve().relative_to(project_data_dir.resolve())
    key = f"{rel.as_posix()}|{job.element_name}|{job.complex_type_name}"
    if job.xml_element_name and job.xml_element_name != job.element_name:
        key += f"|xml:{job.xml_element_name}"
    if job.parent_element_name:
        key += f"|parent:{job.parent_element_name}"
    return key


class NestedImportEngine:
    """Kapselt die rekursive Erstellung und Befuellung verschachtelter Tabellen."""

    def __init__(
        self,
        conn: sqlite3.Connection,
        idx: XsdIndex,
        key_manager: Optional[KeySelectionManager] = None,
        import_key: str = "",
    ) -> None:
        self.conn = conn
        self.idx = idx
        self.key_manager = key_manager
        self.import_key = import_key
        self.nested_tables: Dict[str, NestedTableState] = {}

    @staticmethod
    def _type_map(columns: List[Tuple[str, str]]) -> Dict[str, str]:
        return {col: sql_type for col, sql_type in columns}

    def _choose_nested_pk(
        self,
        state_key: str,
        options: List[str],
        prompt: str,
        auto_column_name: str,
    ) -> str:
        if self.key_manager:
            return self.key_manager.choose_nested_pk(
                self.import_key,
                state_key,
                options,
                prompt,
                auto_column_name,
            )
        return prompt_pk_choice(prompt, options, auto_column_name)

    def _infer_leaf_columns(self, xml_element: ET.Element) -> List[Tuple[str, str]]:
        """Leitet Spalten aus Blattknoten im XML ab (mit Typheuristik)."""
        inferred: List[Tuple[str, str]] = []
        seen = set()

        for child in list(xml_element):
            tag = localname(child.tag)
            if has_element_children(child):
                continue
            if tag in seen:
                continue
            inferred.append((tag, infer_sql_type_from_text(child.text)))
            seen.add(tag)

        return inferred

    def _columns_from_complex_type(self, complex_type_name: Optional[str]) -> List[Tuple[str, str]]:
        """Erzeugt Spaltenliste aus XSD-complexType (nur Leaf-Felder)."""
        if not complex_type_name:
            return []

        complex_type = self.idx.complex_types.get(complex_type_name)
        if not complex_type:
            return []

        columns: List[Tuple[str, str]] = []
        for col_name, nested_xsd_type in complex_type.elements:
            local_type = self.idx.strip_prefix(nested_xsd_type)

            # Eigene Tabellen fuer komplexe Unterstrukturen.
            if local_type in self.idx.complex_types:
                continue

            builtin, length_hint = self.idx.resolve_to_builtin_base(nested_xsd_type)
            columns.append((col_name, map_xsd_builtin_to_sql(builtin, length_hint)))

        return columns

    def _ensure_additional_columns(
        self,
        state: NestedTableState,
        required_cols: List[str],
        overrides: Optional[Dict[str, str]] = None,
    ) -> None:
        """Fuegt zur Laufzeit neue Spalten hinzu, falls XML neue Felder liefert."""
        existing = {col for col, _ in state.columns}

        for col in required_cols:
            if col in existing:
                continue

            sql_type = overrides.get(col, "TEXT") if overrides else "TEXT"
            self.conn.execute(f'ALTER TABLE "{state.table_name}" ADD COLUMN "{col}" {sql_type}')
            state.columns.append((col, sql_type))
            existing.add(col)

        state.insert_sql = build_insert_sql(state.table_name, [col for col, _ in state.columns])

    def _get_or_create_nested_state(
        self,
        *,
        parent_table: str,
        parent_columns: List[Tuple[str, str]],
        parent_pk: str,
        parent_complex_type: Optional[str],
        child_tag: str,
        child_element: ET.Element,
    ) -> NestedTableState:
        """Erzeugt bei Bedarf eine verschachtelte Tabelle fuer `child_tag`."""
        state_key = f"{parent_table}::{child_tag}"
        table_name = f"{parent_table}__{child_tag}"

        existing = self.nested_tables.get(state_key)
        if existing:
            inferred = self._infer_leaf_columns(child_element)
            self._ensure_additional_columns(
                existing,
                [col for col, _ in inferred],
                {col: sql_type for col, sql_type in inferred},
            )
            return existing

        child_complex_type_name: Optional[str] = None
        xsd_cols: List[Tuple[str, str]] = []

        child_xsd_type = self.idx.get_child_element_type(parent_complex_type, child_tag)
        if child_xsd_type:
            local_child_type = self.idx.strip_prefix(child_xsd_type)
            if local_child_type in self.idx.complex_types:
                child_complex_type_name = local_child_type
                xsd_cols = self._columns_from_complex_type(child_complex_type_name)

        inferred_cols = self._infer_leaf_columns(child_element)
        child_columns = merge_columns(xsd_cols, inferred_cols)
        if not child_columns:
            child_columns = [("value", "TEXT")]

        child_pk_choice = self._choose_nested_pk(
            state_key=state_key,
            options=[col for col, _ in child_columns],
            prompt=f'Primary Key fuer Tabelle "{table_name}" waehlen:',
            auto_column_name=suggest_auto_pk_column(
                f"{table_name}_id",
                [col for col, _ in child_columns],
            ),
        )
        child_pk, child_pk_autoincrement = decode_pk_choice(child_pk_choice)
        if child_pk_autoincrement and child_pk not in {col for col, _ in child_columns}:
            child_columns.append((child_pk, "INTEGER"))

        fk_ref_col = parent_pk
        print(
            f'FK-Referenzspalte aus Eltern-Tabelle "{parent_table}" fuer "{table_name}" '
            f"automatisch: {fk_ref_col}"
        )

        # Child-Tabelle bleibt dedupliziert und enthaelt keine Parent-FK-Spalte.
        # Parent<->Child wird ueber eine eigene Link-Tabelle modelliert (m:n-faehig).
        child_ddl = create_table_sql(
            table_name=table_name,
            columns=child_columns,
            pk_col=child_pk,
            pk_autoincrement=child_pk_autoincrement,
            fks=None,
        )
        print("\nDDL:\n", child_ddl)
        self.conn.execute(child_ddl)
        ensure_table_columns(self.conn, table_name, child_columns)
        ensure_common_indexes(self.conn, table_name, child_columns)

        link_table_name: Optional[str] = None
        link_parent_col: Optional[str] = None
        link_child_col: Optional[str] = None
        link_insert_sql: Optional[str] = None

        if fk_ref_col:
            parent_types = self._type_map(parent_columns)
            child_types = self._type_map(child_columns)
            link_table_name = f"{table_name}__link"
            link_parent_col = f"{parent_table}_{fk_ref_col}"
            link_child_col = f"{table_name}_{child_pk}"

            link_ddl = create_link_table_sql(
                link_table_name=link_table_name,
                parent_col=link_parent_col,
                parent_type=parent_types.get(fk_ref_col, "TEXT"),
                parent_table=parent_table,
                parent_ref_col=fk_ref_col,
                child_col=link_child_col,
                child_type=child_types.get(child_pk, "TEXT"),
                child_table=table_name,
                child_ref_col=child_pk,
            )
            print("\nDDL:\n", link_ddl)
            self.conn.execute(link_ddl)
            link_insert_sql = build_link_insert_sql(
                table_name=link_table_name,
                parent_col=link_parent_col,
                child_col=link_child_col,
            )

        state = NestedTableState(
            table_name=table_name,
            columns=child_columns,
            pk_col=child_pk,
            pk_autoincrement=child_pk_autoincrement,
            parent_ref_col=fk_ref_col,
            complex_type_name=child_complex_type_name,
            insert_sql=build_insert_sql(table_name, [col for col, _ in child_columns]),
            link_table_name=link_table_name,
            link_parent_col=link_parent_col,
            link_child_col=link_child_col,
            link_insert_sql=link_insert_sql,
        )
        self.nested_tables[state_key] = state
        return state

    def insert_nested_rows(
        self,
        *,
        parent_element: ET.Element,
        parent_table: str,
        parent_columns: List[Tuple[str, str]],
        parent_pk: str,
        parent_row: Dict[str, object],
        parent_complex_type: Optional[str],
    ) -> None:
        """Durchlaeuft rekursiv alle Child-Knoten und schreibt sie in Untertabellen."""
        for child in list(parent_element):
            if not has_element_children(child):
                continue

            child_tag = localname(child.tag)
            state = self._get_or_create_nested_state(
                parent_table=parent_table,
                parent_columns=parent_columns,
                parent_pk=parent_pk,
                parent_complex_type=parent_complex_type,
                child_tag=child_tag,
                child_element=child,
            )

            row = {col: None for col, _ in state.columns}
            new_cols: List[str] = []
            existing_cols = {col for col, _ in state.columns}

            for sub in list(child):
                if has_element_children(sub):
                    continue
                tag = localname(sub.tag)
                row[tag] = (sub.text or "").strip()
                if tag not in existing_cols:
                    new_cols.append(tag)

            if new_cols:
                self._ensure_additional_columns(state, new_cols)
                for col in new_cols:
                    row.setdefault(col, None)

            try:
                cursor = self.conn.execute(state.insert_sql, [row.get(col) for col, _ in state.columns])
                if state.pk_col and state.pk_autoincrement:
                    row[state.pk_col] = cursor.lastrowid
                state.inserted_rows += 1
            except sqlite3.IntegrityError as exc:
                if "UNIQUE constraint failed" in str(exc):
                    state.skipped_duplicates += 1
                else:
                    raise

            if state.link_insert_sql and state.parent_ref_col and state.pk_col:
                parent_value = parent_row.get(state.parent_ref_col)
                child_value = row.get(state.pk_col)
                if parent_value is not None and child_value is not None:
                    before_changes = self.conn.total_changes
                    self.conn.execute(state.link_insert_sql, (parent_value, child_value))
                    if self.conn.total_changes > before_changes:
                        state.linked_rows += 1

            self.insert_nested_rows(
                parent_element=child,
                parent_table=state.table_name,
                parent_columns=state.columns,
                parent_pk=state.pk_col,
                parent_row=row,
                parent_complex_type=state.complex_type_name,
            )


# ----------------------------
# Public API
# ----------------------------


def import_xml_blocks_to_sqlite(
    xml_path: Path,
    xsd_main: Path,
    xsd_defs: Path,
    sqlite_path: Path,
    element_name: str,
    complex_type_name: Optional[str] = None,
    xml_element_name: Optional[str] = None,
    parent_element_name: Optional[str] = None,
    key_manager: Optional[KeySelectionManager] = None,
    import_key: Optional[str] = None,
) -> None:
    """Importiert alle XML-Bloecke `element_name` inkl. verschachtelter Child-Knoten."""

    idx = XsdIndex()
    idx.load_xsd(xsd_main)
    if xsd_defs.resolve() != xsd_main.resolve():
        idx.load_xsd(xsd_defs)

    ct_name = complex_type_name
    if not ct_name:
        guessed = f"{element_name}_type"
        if guessed in idx.complex_types:
            ct_name = guessed
        elif element_name in idx.element_to_type:
            ct_name = idx.strip_prefix(idx.element_to_type[element_name])
        else:
            raise ValueError(f"Kein complexType gefunden/ableitbar fuer {element_name}")

    ct = idx.complex_types.get(ct_name)
    if not ct:
        raise ValueError(f"complexType '{ct_name}' nicht im XSD-Index gefunden.")

    top_columns: List[Tuple[str, str]] = []
    for col_name, xsd_type in ct.elements:
        builtin, length_hint = idx.resolve_to_builtin_base(xsd_type)
        top_columns.append((col_name, map_xsd_builtin_to_sql(builtin, length_hint)))

    top_options = [col for col, _ in top_columns]
    top_auto_pk_col = suggest_auto_pk_column(
        f"{element_name}_id",
        top_options,
    )
    if key_manager and import_key:
        top_pk_choice = key_manager.choose_top_pk(
            import_key=import_key,
            options=top_options,
            prompt=f'Primary Key fuer Tabelle "{element_name}" waehlen:',
            auto_column_name=top_auto_pk_col,
        )
    else:
        top_pk_choice = prompt_pk_choice(
            f'Primary Key fuer Tabelle "{element_name}" waehlen:',
            top_options,
            top_auto_pk_col,
        )
    top_pk, top_pk_autoincrement = decode_pk_choice(top_pk_choice)
    if top_pk_autoincrement and top_pk not in {col for col, _ in top_columns}:
        top_columns.append((top_pk, "INTEGER"))

    top_ddl = create_table_sql(
        table_name=element_name,
        columns=top_columns,
        pk_col=top_pk,
        pk_autoincrement=top_pk_autoincrement,
    )
    print("\nDDL:\n", top_ddl)

    conn = sqlite3.connect(sqlite_path)
    engine = NestedImportEngine(
        conn,
        idx,
        key_manager=key_manager,
        import_key=import_key or element_name,
    )

    try:
        conn.execute("PRAGMA foreign_keys = ON")
        conn.execute(top_ddl)
        ensure_table_columns(conn, element_name, top_columns)
        ensure_common_indexes(conn, element_name, top_columns)

        root = ET.parse(xml_path).getroot()
        search_name = xml_element_name or element_name
        blocks = [element for element in root.iter() if localname(element.tag) == search_name]

        if parent_element_name:
            parent_map = {child: parent for parent in root.iter() for child in list(parent)}
            blocks = [
                element
                for element in blocks
                if localname(parent_map[element].tag) == parent_element_name
            ]

        if not blocks:
            if parent_element_name:
                raise ValueError(
                    f'Keine <{search_name}>-Bloecke unter <{parent_element_name}> gefunden.'
                )
            raise ValueError(f"Keine <{search_name}>-Bloecke gefunden.")

        col_names = [col for col, _ in top_columns]
        top_insert_sql = build_insert_sql(element_name, col_names)

        top_duplicates = 0
        for block in blocks:
            top_row = {col: None for col in col_names}
            for child in list(block):
                tag = localname(child.tag)
                if tag in top_row:
                    top_row[tag] = (child.text or "").strip()
            for attr_name, attr_value in block.attrib.items():
                if attr_name in top_row:
                    top_row[attr_name] = (attr_value or "").strip()
            if "value" in top_row:
                text_value = (block.text or "").strip()
                if text_value:
                    top_row["value"] = text_value

            try:
                cursor = conn.execute(top_insert_sql, [top_row[col] for col in col_names])
                if top_pk_autoincrement:
                    top_row[top_pk] = cursor.lastrowid
                    # Falls die Tabelle bereits ohne echte AUTOINCREMENT-PK existierte,
                    # schreiben wir den generierten Wert explizit in die ID-Spalte.
                    conn.execute(
                        f'UPDATE "{element_name}" SET "{top_pk}" = ? '
                        f'WHERE rowid = ? AND "{top_pk}" IS NULL',
                        (cursor.lastrowid, cursor.lastrowid),
                    )
            except sqlite3.IntegrityError as exc:
                if "UNIQUE constraint failed" in str(exc):
                    top_duplicates += 1
                else:
                    raise

            engine.insert_nested_rows(
                parent_element=block,
                parent_table=element_name,
                parent_columns=top_columns,
                parent_pk=top_pk,
                parent_row=top_row,
                parent_complex_type=ct_name,
            )

        conn.commit()

        inserted_top = len(blocks) - top_duplicates
        print(f"\nImport fertig. Eingefuegte Zeilen in {element_name}: {inserted_top}")
        if top_duplicates:
            print(f"Uebersprungene Duplikate in {element_name} (PK-Kollision): {top_duplicates}")

        if engine.nested_tables:
            nested_total = sum(state.inserted_rows for state in engine.nested_tables.values())
            print(f"Eingefuegte verschachtelte Zeilen: {nested_total}")
            for state in engine.nested_tables.values():
                message = f"  {state.table_name}: {state.inserted_rows}"
                if state.skipped_duplicates:
                    message += f" (Duplikate uebersprungen: {state.skipped_duplicates})"
                if state.link_table_name:
                    message += f", Link-Zeilen: {state.linked_rows}"
                print(message)
    finally:
        conn.close()


def _run_import_jobs(
    *,
    jobs: List[ImportJob],
    project_data_dir: Path,
    xsd_defs: Path,
    sqlite_path: Path,
    key_manager: KeySelectionManager,
) -> Tuple[int, int]:
    """Fuehrt eine gegebene Jobliste aus und liefert (erfolgreich, uebersprungen)."""
    print(f"\nGefundene Jobs: {len(jobs)}")
    ok = 0
    skipped = 0

    for idx, job in enumerate(jobs, start=1):
        import_key = build_import_key(project_data_dir, job)
        source_name = job.xml_element_name or job.element_name
        if job.parent_element_name:
            source_name = f"{job.parent_element_name}/{source_name}"
        print("\n" + "=" * 72)
        print(
            f"Job {idx}/{len(jobs)}: {job.xml_path.name} -> "
            f'{job.element_name} [XML: {source_name}] ({job.complex_type_name})'
        )
        print(f"Key: {import_key}")

        try:
            import_xml_blocks_to_sqlite(
                xml_path=job.xml_path,
                xsd_main=job.xsd_main,
                xsd_defs=xsd_defs,
                sqlite_path=sqlite_path,
                element_name=job.element_name,
                complex_type_name=job.complex_type_name,
                xml_element_name=job.xml_element_name,
                parent_element_name=job.parent_element_name,
                key_manager=key_manager,
                import_key=import_key,
            )
            ok += 1
        except ValueError as exc:
            skipped += 1
            print(f"UEBERSPRUNGEN: {exc}")

    zkp_stats: Dict[str, int] = {}
    try:
        zkp_stats = sync_zkp_to_database(sqlite_path)
    except Exception as exc:
        print(f"WARNUNG: ZKP-Synchronisierung fehlgeschlagen: {exc}")

    key_manager.save()
    print("\n" + "=" * 72)
    print(f"Fertig. Erfolgreich: {ok}, Uebersprungen: {skipped}")
    if zkp_stats:
        print("ZKP-Aktualisierung (Zeilen):")
        for table_name, changed in sorted(zkp_stats.items()):
            print(f"  {table_name}: {changed}")
    print(f"DB: {sqlite_path}")
    print(f"Key-Config: {key_manager.config_path}")
    return ok, skipped


def configure_single_xml_workflow(
    *,
    project_data_dir: Path,
    schemas_dir: Path,
    xsd_defs: Path,
    sqlite_path: Path,
    key_config_path: Path,
) -> None:
    """Konfiguriert/aktualisiert Keys fuer genau eine ausgewaehlte XML-Datei."""
    all_xml_files = list_project_data_xml_files(project_data_dir)
    if not all_xml_files:
        print("Keine XML-Dateien im project_data-Ordner gefunden.")
        return

    xml_names = [p.name for p in all_xml_files]
    selected_xml = prompt_choice("Welche XML-Datei moechtest du konfigurieren?", xml_names)
    if not selected_xml:
        print("Abgebrochen.")
        return

    key_manager = KeySelectionManager(config_path=key_config_path, mode="configure")
    selected_jobs = discover_jobs(
        project_data_dir,
        schemas_dir,
        key_manager=key_manager,
        allow_manual_xsd_selection=True,
        xml_name_filter={selected_xml},
    )
    if not selected_jobs:
        print(f"Hinweis: {selected_xml} hat keine passenden Importziele in den XSD-Dateien.")
        key_manager.save()
        return

    _run_import_jobs(
        jobs=selected_jobs,
        project_data_dir=project_data_dir,
        xsd_defs=xsd_defs,
        sqlite_path=sqlite_path,
        key_manager=key_manager,
    )


def configure_auto_update_files_workflow(
    *,
    project_data_dir: Path,
    key_config_path: Path,
) -> None:
    """Speichert, welche XML-Dateien bei Auto-Update beruecksichtigt werden."""
    all_xml_files = list_project_data_xml_files(project_data_dir)
    if not all_xml_files:
        print("Keine XML-Dateien im project_data-Ordner gefunden.")
        return

    xml_names = [p.name for p in all_xml_files]
    key_manager = KeySelectionManager(config_path=key_config_path, mode="configure")
    current = key_manager.get_update_files()
    if current:
        print("\nAktuell fuer Auto-Update gespeichert:")
        for name in current:
            print(f"  - {name}")
    else:
        print("\nAktuell ist keine Auto-Update-Dateiliste gespeichert.")

    selected = prompt_multi_file_selection(xml_names)
    key_manager.set_update_files(selected)
    key_manager.save()
    print(f"\nAuto-Update-Dateiliste gespeichert ({len(selected)} Dateien).")


def run_saved_auto_update_workflow(
    *,
    project_data_dir: Path,
    schemas_dir: Path,
    xsd_defs: Path,
    sqlite_path: Path,
    key_config_path: Path,
) -> None:
    """Fuehrt Auto-Update fuer die gespeicherte Dateiliste aus."""
    if not key_config_path.exists():
        print(
            f"Keine Key-Konfiguration gefunden: {key_config_path}. "
            "Bitte zuerst eine XML konfigurieren und die Auto-Update-Dateiliste festlegen."
        )
        return

    key_manager = KeySelectionManager(config_path=key_config_path, mode="update")
    selected_files = key_manager.get_update_files()
    if not selected_files:
        print("Keine Dateien fuer Auto-Update gespeichert. Bitte Dateiliste zuerst festlegen.")
        return

    selected_set = set(selected_files)
    jobs = discover_jobs(
        project_data_dir,
        schemas_dir,
        key_manager=key_manager,
        allow_manual_xsd_selection=True,
        xml_name_filter=selected_set,
    )
    files_with_jobs = {job.xml_path.name for job in jobs}
    missing = sorted(selected_set - files_with_jobs, key=str.lower)
    for name in missing:
        print(f"Hinweis: {name} uebersprungen (keine passenden Importziele in XSD gefunden).")

    if not jobs:
        key_manager.save()
        print("Keine passenden Jobs fuer die gespeicherte Auto-Update-Dateiliste gefunden.")
        return

    _run_import_jobs(
        jobs=jobs,
        project_data_dir=project_data_dir,
        xsd_defs=xsd_defs,
        sqlite_path=sqlite_path,
        key_manager=key_manager,
    )


def run_tool_menu(
    *,
    project_data_dir: Path,
    schemas_dir: Path,
    xsd_defs: Path,
    sqlite_path: Path,
    key_config_path: Path,
) -> None:
    """Interaktives Hauptmenue fuer Konfiguration und Auto-Update."""
    while True:
        action = prompt_required_choice(
            "\nAktion waehlen:",
            [
                "Einzelne XML waehlen und Keys festlegen",
                "Dateiliste fuer automatisierten DB-Update festlegen",
                "Automatisierten DB-Update ausfuehren",
                "Beenden",
            ],
        )

        if action == "Einzelne XML waehlen und Keys festlegen":
            configure_single_xml_workflow(
                project_data_dir=project_data_dir,
                schemas_dir=schemas_dir,
                xsd_defs=xsd_defs,
                sqlite_path=sqlite_path,
                key_config_path=key_config_path,
            )
            continue

        if action == "Dateiliste fuer automatisierten DB-Update festlegen":
            configure_auto_update_files_workflow(
                project_data_dir=project_data_dir,
                key_config_path=key_config_path,
            )
            continue

        if action == "Automatisierten DB-Update ausfuehren":
            run_saved_auto_update_workflow(
                project_data_dir=project_data_dir,
                schemas_dir=schemas_dir,
                xsd_defs=xsd_defs,
                sqlite_path=sqlite_path,
                key_config_path=key_config_path,
            )
            continue

        print("Beendet.")
        return


if __name__ == "__main__":
    run_tool_menu(
        project_data_dir=Path(
            r"BVNG_Advanced_A1_Var1_Ed_01_120\BVNG_Advanced_A1_Var1_Ed_01_120\STBD_Outside_(Shaft1)_SPU_SW_Image_26.02.2025\Node_1\project_data"
        ),
        schemas_dir=Path(r"stylesheets\schemas"),
        xsd_defs=Path(r"stylesheets\schemas\MCS6_Definitions.xsd"),
        sqlite_path=Path("out.db"),
        key_config_path=Path(r"importers\key_selection.json"),
    )

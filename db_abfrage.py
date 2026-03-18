"""Kleine Hilfsfunktionen fuer Datenbankabfragen und ZKP-Updates auf SQLite."""

from __future__ import annotations

import csv
import re
import sqlite3
from pathlib import Path
from typing import Any, Sequence

DB_PATH = Path("out.db")
ZKP_SOURCE_PATH = Path("Ed_01_10_2")
ZKP_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xls"}
TABLE_NAME = "ITEM"
LIMIT = 10
SEARCH_LIMIT_PER_TABLE = 10


def execute_query(
    db_path: Path | str,
    query: str,
    params: Sequence[Any] = (),
) -> list[dict[str, Any]]:
    """Fuehrt eine SQL-Abfrage aus und liefert Ergebniszeilen als Dict-Liste.

    Hinweise:
    - Fuer Benutzereingaben immer Platzhalter verwenden, z. B. "... WHERE id = ?".
    - `params` wird dann als Tupel/Liste uebergeben.
    """
    path = Path(db_path)
    if not path.exists():
        raise FileNotFoundError(f"SQLite-Datei nicht gefunden: {path}")

    with sqlite3.connect(path) as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.execute(query, tuple(params))
        rows = cursor.fetchall()

    return [dict(row) for row in rows]


def list_tables(db_path: Path | str) -> list[str]:
    """Liefert alle Tabellen in der Datenbank (ohne SQLite-Systemtabellen)."""
    rows = execute_query(
        db_path,
        "SELECT name FROM sqlite_master WHERE type = 'table' AND name NOT LIKE 'sqlite_%' ORDER BY name",
    )
    return [row["name"] for row in rows]


def _quote_ident(identifier: str) -> str:
    """Quoted SQL-Identifier fuer SQLite."""
    return f'"{identifier.replace("\"", "\"\"")}"'


def search_database(
    db_path: Path | str,
    search_term: str,
    *,
    tables: Sequence[str] | None = None,
    limit_per_table: int = 20,
    case_sensitive: bool = False,
) -> dict[str, list[dict[str, Any]]]:
    """Durchsucht die DB tabellenuebergreifend nach einem Suchbegriff.

    Es werden alle Spalten einer Tabelle per CAST(... AS TEXT) durchsucht.
    Rueckgabe: Dict mit Tabellenname -> Trefferzeilen.
    """
    path = Path(db_path)
    if not path.exists():
        raise FileNotFoundError(f"SQLite-Datei nicht gefunden: {path}")
    if not search_term:
        raise ValueError("search_term darf nicht leer sein.")
    if limit_per_table <= 0:
        raise ValueError("limit_per_table muss > 0 sein.")

    try:
        with sqlite3.connect(path) as conn:
            conn.row_factory = sqlite3.Row

            if tables is None:
                selected_tables = [
                    row["name"]
                    for row in conn.execute(
                        "SELECT name FROM sqlite_master "
                        "WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
                    )
                ]
            else:
                selected_tables = list(tables)

            result: dict[str, list[dict[str, Any]]] = {}
            needle = f"%{search_term}%"

            for table_name in selected_tables:
                columns = [row[1] for row in conn.execute(f'PRAGMA table_info({_quote_ident(table_name)})')]
                if not columns:
                    continue

                if case_sensitive:
                    where_parts = [f'CAST({_quote_ident(col)} AS TEXT) LIKE ?' for col in columns]
                    params: list[Any] = [needle] * len(columns) + [limit_per_table]
                else:
                    where_parts = [f'LOWER(CAST({_quote_ident(col)} AS TEXT)) LIKE LOWER(?)' for col in columns]
                    params = [needle] * len(columns) + [limit_per_table]

                query = (
                    f'SELECT * FROM {_quote_ident(table_name)} '
                    f'WHERE {" OR ".join(where_parts)} '
                    "LIMIT ?"
                )

                rows = conn.execute(query, tuple(params)).fetchall()
                if rows:
                    result[table_name] = [dict(row) for row in rows]

        return result
    except sqlite3.OperationalError as exc:
        raise RuntimeError(
            "DB-Suche fehlgeschlagen. Bitte sicherstellen, dass die DB-Datei nicht gesperrt ist "
            "(z. B. durch DB Browser/Editor) und Schreib-/Leserechte vorhanden sind."
        ) from exc


def _normalize_zkp(raw_value: str) -> str:
    """Entfernt Zeilenumbrueche/Mehrfachspaces aus ZKP-Texten."""
    return " ".join(raw_value.split())


def _normalize_sys_id(raw_value: object) -> str:
    """Normalisiert SYS_ID-Werte fuer robuste Vergleiche."""
    return str(raw_value).strip() if raw_value is not None else ""


def _canonicalize_sys_id_for_zkp(sys_id: str) -> str:
    """Kanonische SYS_ID mit letztem Segment = 0."""
    value = _normalize_sys_id(sys_id)
    match = re.match(r"^(.*\.)(\d+)$", value)
    if not match:
        return value
    return f"{match.group(1)}0"


def _build_canonical_zkp_mapping(mapping: dict[str, str]) -> dict[str, str]:
    """Baut eine konfliktfreie kanonische SYS_ID->ZKP-Zuordnung."""
    canonical: dict[str, str] = {}
    conflicts: set[str] = set()

    for sys_id, zkp in mapping.items():
        key = _canonicalize_sys_id_for_zkp(sys_id)
        existing = canonical.get(key)
        if existing is None:
            canonical[key] = zkp
            continue
        if existing != zkp:
            conflicts.add(key)

    for key in conflicts:
        canonical.pop(key, None)
    return canonical


def _sys_id_prefix3(sys_id: str) -> str:
    """Liefert den SYS_ID-Praefix aus den ersten 3 Gruppen."""
    value = _normalize_sys_id(sys_id)
    if not value:
        return ""
    parts = value.split(".")
    if len(parts) < 3:
        return ""
    return ".".join(parts[:3])


def _build_prefix3_zkp_mapping(mapping: dict[str, str]) -> dict[str, str]:
    """Erzeugt konfliktfreie ZKP-Zuordnung ueber SYS_ID-Praefix (erste 3 Gruppen)."""
    prefix_map: dict[str, str] = {}
    conflicts: set[str] = set()

    for sys_id, zkp in mapping.items():
        prefix = _sys_id_prefix3(sys_id)
        if not prefix:
            continue
        existing = prefix_map.get(prefix)
        if existing is None:
            prefix_map[prefix] = zkp
            continue
        if existing != zkp:
            conflicts.add(prefix)

    for prefix in conflicts:
        prefix_map.pop(prefix, None)
    return prefix_map


def _lookup_zkp_for_sys_id(
    sys_id: str,
    exact_mapping: dict[str, str],
    canonical_mapping: dict[str, str],
    prefix3_mapping: dict[str, str],
) -> str | None:
    """Sucht ZKP fuer eine SYS_ID mit Fallback-Kaskade."""
    value = _normalize_sys_id(sys_id)
    if not value:
        return None

    exact = exact_mapping.get(value)
    if exact:
        return exact

    canonical = canonical_mapping.get(_canonicalize_sys_id_for_zkp(value))
    if canonical:
        return canonical

    return prefix3_mapping.get(_sys_id_prefix3(value))


def _parse_sysid_zkp_rows(rows: list[list[object]]) -> dict[str, str]:
    """Extrahiert SYS_ID -> ZKP aus tabellarischen Zeilen."""
    def normalize_header(cell: object) -> str:
        return re.sub(r"[^A-Z0-9]", "", str(cell).upper()) if cell is not None else ""

    def score_sys_header(cell: object) -> int:
        norm = normalize_header(cell)
        if norm == "SYSID":
            return 4
        if norm.startswith("SYSID"):
            return 3
        if "SYSID" in norm:
            return 2
        if norm.startswith("SYS") and "ID" in norm:
            return 1
        return 0

    def score_zkp_header(cell: object) -> int:
        norm = normalize_header(cell)
        if norm == "ZKP":
            return 4
        if norm.startswith("ZKP") or norm.endswith("ZKP"):
            return 3
        if "ZKP" in norm:
            return 2
        return 0

    def looks_like_sys_id(value: str) -> bool:
        parts = value.split(".")
        if len(parts) < 3:
            return False
        return all(part.isdigit() for part in parts)

    def is_valid_zkp_cell(raw_value: object) -> bool:
        if raw_value is None:
            return False
        text = str(raw_value).strip()
        if not text:
            return False
        upper = text.upper()
        return upper not in {"#N/A", "N/A", "NA", "NULL", "-"}

    def count_sys_hits(start_idx: int, col_idx: int, window: int = 250) -> int:
        end = min(len(rows), start_idx + 1 + window)
        hits = 0
        for r in rows[start_idx + 1 : end]:
            if col_idx >= len(r):
                continue
            value = _normalize_sys_id(r[col_idx])
            if looks_like_sys_id(value):
                hits += 1
        return hits

    def count_zkp_hits(start_idx: int, col_idx: int, window: int = 250) -> int:
        end = min(len(rows), start_idx + 1 + window)
        hits = 0
        for r in rows[start_idx + 1 : end]:
            if col_idx >= len(r):
                continue
            if is_valid_zkp_cell(r[col_idx]):
                hits += 1
        return hits

    header_idx = None
    sys_col = -1
    zkp_col = -1
    best_score = -1

    for idx, row in enumerate(rows):
        sys_candidates = [(col_idx, score_sys_header(cell)) for col_idx, cell in enumerate(row)]
        sys_candidates = [item for item in sys_candidates if item[1] > 0]
        zkp_candidates = [(col_idx, score_zkp_header(cell)) for col_idx, cell in enumerate(row)]
        zkp_candidates = [item for item in zkp_candidates if item[1] > 0]
        if not sys_candidates or not zkp_candidates:
            continue

        sys_ranked = [
            (col_idx, hdr_score, count_sys_hits(idx, col_idx))
            for col_idx, hdr_score in sys_candidates
        ]
        sys_ranked.sort(key=lambda item: (item[2], item[1]), reverse=True)
        best_sys_col, best_sys_hdr_score, best_sys_hits = sys_ranked[0]
        if best_sys_hits == 0:
            continue

        zkp_ranked = [
            (col_idx, hdr_score, count_zkp_hits(idx, col_idx))
            for col_idx, hdr_score in zkp_candidates
        ]
        zkp_ranked.sort(key=lambda item: (item[2], item[1]), reverse=True)
        best_zkp_col, best_zkp_hdr_score, best_zkp_hits = zkp_ranked[0]

        score = best_sys_hits * 1000 + best_zkp_hits * 10 + best_sys_hdr_score + best_zkp_hdr_score
        if score > best_score:
            best_score = score
            header_idx = idx
            sys_col = best_sys_col
            zkp_col = best_zkp_col

    if header_idx is None or sys_col < 0 or zkp_col < 0:
        return {}

    mapping: dict[str, str] = {}
    for row in rows[header_idx + 1 :]:
        if sys_col >= len(row):
            continue
        sys_id = _normalize_sys_id(row[sys_col])
        if not sys_id:
            continue
        if not looks_like_sys_id(sys_id):
            continue

        raw_cell = row[zkp_col] if zkp_col < len(row) else None
        if not is_valid_zkp_cell(raw_cell):
            continue
        raw_zkp = str(raw_cell).strip()
        if not raw_zkp:
            continue

        zkp = _normalize_zkp(raw_zkp)
        if zkp and sys_id not in mapping:
            mapping[sys_id] = zkp

    return mapping


def _load_sysid_zkp_from_csv(csv_path: Path) -> dict[str, str]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.reader(fh, delimiter=";"))
    return _parse_sysid_zkp_rows([[cell for cell in row] for row in rows])


def _load_sysid_zkp_from_excel(excel_path: Path) -> dict[str, str]:
    suffix = excel_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        try:
            import openpyxl  # type: ignore
        except ModuleNotFoundError:
            return {}

        workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        mapping: dict[str, str] = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            part = _parse_sysid_zkp_rows(rows)
            for sys_id, zkp in part.items():
                if sys_id not in mapping:
                    mapping[sys_id] = zkp
        return mapping

    if suffix == ".xls":
        try:
            import xlrd  # type: ignore
        except ModuleNotFoundError:
            return {}

        workbook = xlrd.open_workbook(excel_path)
        mapping = {}
        for sheet in workbook.sheets():
            rows = [sheet.row_values(idx) for idx in range(sheet.nrows)]
            part = _parse_sysid_zkp_rows(rows)
            for sys_id, zkp in part.items():
                if sys_id not in mapping:
                    mapping[sys_id] = zkp
        return mapping

    return {}


def _collect_zkp_source_files(source_path: Path) -> list[Path]:
    if source_path.is_dir():
        return sorted(
            [
                path
                for path in source_path.iterdir()
                if path.is_file()
                and path.suffix.lower() in ZKP_EXCEL_SUFFIXES
                and not path.name.startswith("~$")
            ],
            key=lambda path: path.name.lower(),
        )

    if source_path.is_file():
        if source_path.suffix.lower() in ZKP_EXCEL_SUFFIXES or source_path.suffix.lower() == ".csv":
            return [source_path]
    return []


def load_sysid_zkp_mapping(source_path: Path | str = ZKP_SOURCE_PATH) -> tuple[dict[str, str], list[str]]:
    """Liest SYS_ID -> ZKP aus allen Excel-Dateien im Ordner (CSV-Fallback pro Datei)."""
    source = Path(source_path)
    mapping: dict[str, str] = {}
    used_sources: list[str] = []

    for source_file in _collect_zkp_source_files(source):
        part: dict[str, str] = {}
        used_file: Path | None = None

        try:
            if source_file.suffix.lower() == ".csv":
                part = _load_sysid_zkp_from_csv(source_file)
                used_file = source_file if part else None
            else:
                part = _load_sysid_zkp_from_excel(source_file)
                used_file = source_file if part else None
        except Exception:
            part = {}
            used_file = None

        if not part and source_file.suffix.lower() in ZKP_EXCEL_SUFFIXES:
            csv_fallback = source_file.with_suffix(".csv")
            if csv_fallback.exists():
                try:
                    part = _load_sysid_zkp_from_csv(csv_fallback)
                    used_file = csv_fallback if part else None
                except Exception:
                    part = {}
                    used_file = None

        if not part:
            continue

        for sys_id, zkp in part.items():
            if sys_id not in mapping:
                mapping[sys_id] = zkp

        if used_file is not None:
            used_sources.append(str(used_file))

    return mapping, used_sources


def _get_sysid_columns_by_table(conn: sqlite3.Connection) -> list[tuple[str, list[str]]]:
    """Liefert Tabellen plus erkannte SYS_ID-Spalten."""
    tables = [row[0] for row in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
    )]

    result: list[tuple[str, list[str]]] = []
    for table in tables:
        table_info = conn.execute(f'PRAGMA table_info("{table}")').fetchall()
        ordered_cols = [row[1] for row in table_info]
        sys_cols = [
            col
            for col in ordered_cols
            if "SYS_ID" in col.upper() or col == "SysID"
        ]
        if sys_cols:
            result.append((table, sys_cols))
    return result


def update_zkp_from_csv(
    db_path: Path | str,
    csv_path: Path | str = ZKP_SOURCE_PATH,
) -> dict[str, int]:
    """Schreibt ZKP-Werte anhand SYS_ID in alle passenden DB-Tabellen."""
    mapping, used_sources = load_sysid_zkp_mapping(csv_path)
    canonical_mapping = _build_canonical_zkp_mapping(mapping)
    prefix3_mapping = _build_prefix3_zkp_mapping(mapping)
    if not mapping:
        raise ValueError("Keine SYS_ID->ZKP-Zuordnungen aus den ZKP-Quellen gelesen.")

    print(
        "ZKP-Quellen geladen: "
        f"{len(used_sources)} Datei(en)"
    )

    path = Path(db_path)
    if not path.exists():
        raise FileNotFoundError(f"SQLite-Datei nicht gefunden: {path}")

    updates: dict[str, int] = {}
    with sqlite3.connect(path) as conn:
        conn.execute("PRAGMA journal_mode=MEMORY")
        conn.execute("PRAGMA synchronous=NORMAL")
        sysid_tables = _get_sysid_columns_by_table(conn)

        for table, sys_cols in sysid_tables:
            col_names = [row[1] for row in conn.execute(f'PRAGMA table_info("{table}")')]
            if "ZKP" not in col_names:
                conn.execute(f'ALTER TABLE "{table}" ADD COLUMN "ZKP" TEXT')

            # Alte Zuordnungen loeschen, damit fehlende Eintraege leer bleiben.
            conn.execute(f'UPDATE "{table}" SET "ZKP" = NULL')

            before = conn.total_changes
            select_cols_sql = ", ".join(f'"{col}"' for col in sys_cols)
            row_data = conn.execute(
                f'SELECT rowid, {select_cols_sql} FROM "{table}"'
            ).fetchall()
            fallback_updates: list[tuple[str, int]] = []
            for row in row_data:
                rowid = row[0]
                candidate_sys_ids = row[1:]
                zkp: str | None = None
                for raw_sys_id in candidate_sys_ids:
                    sys_id = _normalize_sys_id(raw_sys_id)
                    if not sys_id:
                        continue
                    zkp = _lookup_zkp_for_sys_id(
                        sys_id,
                        mapping,
                        canonical_mapping,
                        prefix3_mapping,
                    )
                    if zkp:
                        break
                if zkp:
                    fallback_updates.append((zkp, rowid))

            if fallback_updates:
                conn.executemany(
                    f'UPDATE "{table}" SET "ZKP" = ? WHERE rowid = ?',
                    fallback_updates,
                )

            updates[table] = conn.total_changes - before

        conn.commit()

    return updates


def main() -> None:
    print("Interaktive DB-Suche")
    print(f"DB: {DB_PATH}")
    print("Leere Eingabe beendet das Programm.\n")

    available_tables = list_tables(DB_PATH)
    print(f"Tabellen ({len(available_tables)}):")
    for table_name in available_tables:
        print(f"  - {table_name}")

    while True:
        term = input("\nSuchbegriff: ").strip()
        if not term:
            print("Beendet.")
            return

        try:
            hits = search_database(
                DB_PATH,
                term,
                limit_per_table=SEARCH_LIMIT_PER_TABLE,
            )
        except RuntimeError as exc:
            print(f"Fehler: {exc}")
            return

        if not hits:
            print("Keine Treffer gefunden.")
            continue

        print(f"Treffer in {len(hits)} Tabellen:")
        for table_name, rows in hits.items():
            print(f"\n[{table_name}] {len(rows)} Treffer")
            for row in rows:
                print(row)


if __name__ == "__main__":
    main()
